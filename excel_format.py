from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='B9A6F5', end_color='B9A6F5', fill_type='solid')
HEADER_FONT = Font(color='1A1530', bold=True, size=11, name='Calibri')
BAND_FILL = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
THIN_BORDER = Border(bottom=Side(style='thin', color='E5E0F5'))

CURRENCY_HINTS = ('raised', 'valuation', 'size', '($m)')
WRAP_HINTS = ('description', 'notes', 'competitors', 'investors', 'rationale', 'thesis', 'highlight', 'how it', 'value created')

# Status -> soft pastel fill, for any column literally named "Status"
STATUS_COLORS = {
    'Strong Candidate': 'C9F2D8', 'Shortlisted Finalist': 'B9A6F5', 'Under Review': 'FDE9C6',
    'Passed': 'F5D2D2', 'On Hold': 'E5E0F5',
}


def style_workbook(ws, df, sheet_title='Data'):
    ws.title = sheet_title[:31]
    n_rows, n_cols = df.shape

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 26

    wrap_cols = set()
    currency_cols = set()
    status_col = None
    for col_idx, col_name in enumerate(df.columns, start=1):
        name_l = str(col_name).lower()
        if any(h in name_l for h in WRAP_HINTS):
            wrap_cols.add(col_idx)
        if any(h in name_l for h in CURRENCY_HINTS):
            currency_cols.add(col_idx)
        if name_l == 'status':
            status_col = col_idx

    # Column widths (capped) + number formats + wrap
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_idx in wrap_cols:
            ws.column_dimensions[letter].width = 48
        else:
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str).head(200)])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 28)

    # Data rows
    for row_idx in range(2, n_rows + 2):
        is_band = (row_idx % 2 == 0)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if col_idx in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')
            if col_idx in currency_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.0,,"M"' if abs(cell.value) > 1000 else '"$"#,##0.0"M"'
            if is_band:
                cell.fill = BAND_FILL
            if status_col and col_idx == status_col and cell.value in STATUS_COLORS:
                cell.fill = PatternFill(start_color=STATUS_COLORS[cell.value], end_color=STATUS_COLORS[cell.value], fill_type='solid')
        if any(c in wrap_cols for c in range(1, n_cols + 1)):
            ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
