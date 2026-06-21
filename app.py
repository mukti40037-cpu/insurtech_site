import sqlite3
import os
import json
import datetime
from functools import wraps
from flask import Flask, jsonify, request, send_from_directory, g
from prypco_data import load_prypco_data

ROOT = os.path.dirname(__file__)
_SIBLING_DB_PATH = os.path.join(ROOT, '..', 'insurtech_dashboard', 'insurtech.db')
_BUNDLED_DB_PATH = os.path.join(ROOT, 'data', 'insurtech.db')
# Local dev keeps using the sibling insurtech_dashboard/ folder if present; a deployed
# copy (no sibling folder available) falls back to the bundled copy in data/.
DB_PATH = os.environ.get('DB_PATH') or (_SIBLING_DB_PATH if os.path.exists(_SIBLING_DB_PATH) else _BUNDLED_DB_PATH)
CACHE_TTL_HOURS = 24

app = Flask(__name__, static_folder=None)
_PRYPCO_CACHE = None

# Viewing this site is open to anyone with the link. Editing (adding/deleting/changing
# companies, custom fields, dashboard widgets, scorecards, etc.) is gated behind a shared
# edit code so the owner can hand it out to specific collaborators without making write
# access fully public. Set the EDIT_ACCESS_TOKEN env var to turn the gate on (e.g. in
# production); when it's unset (local dev), all edits are allowed with no token needed.
EDIT_ACCESS_TOKEN = os.environ.get('EDIT_ACCESS_TOKEN')


def require_edit_access(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if EDIT_ACCESS_TOKEN and request.headers.get('X-Edit-Token') != EDIT_ACCESS_TOKEN:
            return jsonify({'error': 'Editing requires the shared edit code for this site.'}), 401
        return fn(*args, **kwargs)
    return wrapper


@app.route('/api/edit-access/check', methods=['POST'])
def api_edit_access_check():
    payload = request.get_json(force=True) or {}
    token = payload.get('token') or ''
    if not EDIT_ACCESS_TOKEN:
        return jsonify({'ok': True, 'gated': False})
    return jsonify({'ok': token == EDIT_ACCESS_TOKEN, 'gated': True})

EDITABLE_FIELDS = ['Primary Segment', 'Secondary Segment', 'Business Model', 'Geography Region', 'Capital Intensity',
                    'Notes', 'ticker', 'ticker_source', 'ticker_note',
                    'Revenue Model', 'Target Customer', 'Go-to-Market Motion', 'Moat', 'Strategy Notes']

COMPANY_FIELD_MAP = {
    'name': 'Companies', 'parentCompany': 'Parent Company', 'businessStatus': 'Business Status',
    'totalRaised': 'Total Raised', 'lastFinancingValuation': 'Last Financing Valuation',
    'emergingSpaces': 'Emerging Spaces', 'verticals': 'Verticals', 'valuationEstimate': 'Valuation Estimate',
    'successProbability': 'Success Probability', 'competitors': 'Competitors',
    'primaryIndustryCode': 'Primary PitchBook Industry Code', 'lastFinancingDate': 'Last Financing Date',
    'lastFinancingSize': 'Last Financing Size', 'lastFinancingDealType': 'Last Financing Deal Type',
    'activeInvestors': 'Active Investors', 'primaryContact': 'Primary Contact',
    'ownershipStatus': 'Ownership Status', 'companyFinancingStatus': 'Company Financing Status',
    'yearFounded': 'Year Founded', 'hq': 'HQ Location', 'website': 'Website', 'description': 'Description',
    'segment': 'Primary Segment', 'segment2': 'Secondary Segment', 'businessModel': 'Business Model',
    'geo': 'Geography Region', 'capitalIntensity': 'Capital Intensity', 'notes': 'Notes',
    'ticker': 'ticker', 'tickerSource': 'ticker_source', 'tickerNote': 'ticker_note',
    'revenueModel': 'Revenue Model', 'targetCustomer': 'Target Customer',
    'gtm': 'Go-to-Market Motion', 'moat': 'Moat', 'strategyNotes': 'Strategy Notes',
    'realFinPeriod': 'real_fin_period', 'realFinRevenue': 'real_fin_revenue',
    'realFinNetIncome': 'real_fin_net_income', 'realFinKeyRatios': 'real_fin_key_ratios',
    'realFinSource': 'real_fin_source',
}
COMPANY_NUMERIC_FIELDS = {'totalRaised', 'lastFinancingValuation', 'valuationEstimate', 'successProbability',
                          'lastFinancingSize', 'yearFounded'}


def coerce_company_value(json_key, raw):
    if json_key in COMPANY_NUMERIC_FIELDS:
        if raw in (None, ''):
            return None
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None
    if isinstance(raw, str):
        raw = raw.strip()
    return raw if raw not in ('', None) else None


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def ensure_price_cache_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS price_cache (
        ticker TEXT NOT NULL,
        range_key TEXT NOT NULL,
        fetched_at TEXT NOT NULL,
        data TEXT NOT NULL,
        PRIMARY KEY (ticker, range_key)
    )''')
    db.commit()


def ensure_comments_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS comments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        author TEXT,
        category TEXT DEFAULT 'General',
        text TEXT NOT NULL,
        page TEXT,
        created_at TEXT
    )''')
    db.commit()


def ensure_shortlist_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS shortlist (
        company_id TEXT PRIMARY KEY,
        watchlist TEXT DEFAULT 'Default',
        rating INTEGER,
        status TEXT DEFAULT 'Under Review',
        rationale TEXT,
        notes TEXT,
        added_at TEXT
    )''')
    db.commit()


def ensure_score_templates_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS score_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        categories_json TEXT NOT NULL,
        created_at TEXT
    )''')
    db.commit()


def ensure_screening_scorecards_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS screening_scorecards (
        company_id TEXT PRIMARY KEY,
        management_quality INTEGER,
        integration_complexity INTEGER,
        market_timing INTEGER,
        deal_feasibility INTEGER,
        manually_added INTEGER DEFAULT 0,
        updated_at TEXT
    )''')
    db.commit()


def ensure_custom_field_tables(db):
    db.execute('''CREATE TABLE IF NOT EXISTS custom_field_defs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        field_type TEXT NOT NULL DEFAULT 'text',
        created_at TEXT
    )''')
    db.execute('''CREATE TABLE IF NOT EXISTS custom_field_values (
        company_id TEXT NOT NULL,
        field_id INTEGER NOT NULL,
        value TEXT,
        PRIMARY KEY (company_id, field_id)
    )''')
    db.commit()


def ensure_dashboard_widgets_table(db):
    db.execute('''CREATE TABLE IF NOT EXISTS dashboard_widgets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL,
        title TEXT NOT NULL,
        config_json TEXT NOT NULL,
        position INTEGER DEFAULT 0,
        created_at TEXT
    )''')
    db.commit()


US_STATES = {
    'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA','HI','ID','IL','IN','IA','KS','KY','LA','ME','MD',
    'MA','MI','MN','MS','MO','MT','NE','NV','NH','NJ','NM','NY','NC','ND','OH','OK','OR','PA','RI','SC',
    'SD','TN','TX','UT','VT','VA','WA','WV','WI','WY','DC'
}


def parse_country(hq):
    if not hq:
        return None
    parts = [p.strip() for p in hq.split(',')]
    last = parts[-1] if parts else None
    if last and last.upper() in US_STATES:
        return 'United States'
    return last


def row_to_company(d):
    return {
        'id': d.get('Company ID'),
        'name': d.get('Companies'),
        'website': d.get('Website'),
        'description': d.get('Description'),
        'segment': d.get('Primary Segment'),
        'segment2': d.get('Secondary Segment'),
        'businessModel': d.get('Business Model'),
        'geo': d.get('Geography Region'),
        'hq': d.get('HQ Location'),
        'yearFounded': d.get('Year Founded'),
        'capitalIntensity': d.get('Capital Intensity'),
        'notes': d.get('Notes'),
        'revenueModel': d.get('Revenue Model'),
        'targetCustomer': d.get('Target Customer'),
        'gtm': d.get('Go-to-Market Motion'),
        'moat': d.get('Moat'),
        'strategyNotes': d.get('Strategy Notes'),
        'ticker': d.get('ticker'),
        'tickerSource': d.get('ticker_source'),
        'tickerNote': d.get('ticker_note'),
        'totalRaised': d.get('Total Raised'),
        'ownershipStatus': d.get('Ownership Status'),
        'businessStatus': d.get('Business Status'),
        'activeInvestors': d.get('Active Investors'),
        'competitors': d.get('Competitors'),
        'lastEditedAt': d.get('last_edited_at'),
        'isHumanVerified': bool(d.get('is_human_verified')),
        'lastFinancingDate': d.get('Last Financing Date'),
        'lastFinancingSize': d.get('Last Financing Size'),
        'lastFinancingValuation': d.get('Last Financing Valuation'),
        'valuationEstimate': d.get('Valuation Estimate'),
        'lastFinancingDealType': d.get('Last Financing Deal Type'),
        'companyFinancingStatus': d.get('Company Financing Status'),
        'country': parse_country(d.get('HQ Location')),
        'parentCompany': d.get('Parent Company'),
        'primaryContact': d.get('Primary Contact'),
        'successProbability': d.get('Success Probability'),
        'primaryIndustryCode': d.get('Primary PitchBook Industry Code'),
        'emergingSpaces': d.get('Emerging Spaces'),
        'verticals': d.get('Verticals'),
        'realFinPeriod': d.get('real_fin_period'),
        'realFinRevenue': d.get('real_fin_revenue'),
        'realFinNetIncome': d.get('real_fin_net_income'),
        'realFinKeyRatios': d.get('real_fin_key_ratios'),
        'realFinSource': d.get('real_fin_source'),
    }


@app.route('/')
def index():
    return send_from_directory(ROOT, 'index.html')


@app.route('/<path:path>')
def static_files(path):
    return send_from_directory(ROOT, path)


def custom_field_defs_by_id(db):
    ensure_custom_field_tables(db)
    return {r['id']: dict(r) for r in db.execute('SELECT * FROM custom_field_defs')}


def attach_custom_fields(db, companies_list, defs_by_id):
    if not defs_by_id:
        return companies_list
    rows = db.execute('SELECT * FROM custom_field_values')
    by_company = {}
    for r in rows:
        d = defs_by_id.get(r['field_id'])
        if not d:
            continue
        by_company.setdefault(r['company_id'], {})[d['name']] = r['value']
    for c in companies_list:
        c['customFields'] = by_company.get(c['id'], {})
    return companies_list


@app.route('/api/companies')
def api_companies():
    db = get_db()
    rows = db.execute('SELECT * FROM companies').fetchall()
    companies_list = [row_to_company(dict(r)) for r in rows]
    attach_custom_fields(db, companies_list, custom_field_defs_by_id(db))
    return jsonify(companies_list)


@app.route('/api/companies/<company_id>')
def api_company_detail(company_id):
    db = get_db()
    row = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (company_id,)).fetchone()
    if row is None:
        return jsonify({'error': 'not found'}), 404
    c = row_to_company(dict(row))
    attach_custom_fields(db, [c], custom_field_defs_by_id(db))
    return jsonify(c)


@app.route('/api/companies/<company_id>', methods=['POST'])
@require_edit_access
def api_company_edit(company_id):
    db = get_db()
    payload = request.get_json(force=True) or {}

    updates = []
    params = []
    for json_key, col in COMPANY_FIELD_MAP.items():
        if json_key in payload:
            updates.append(f'"{col}" = ?')
            params.append(coerce_company_value(json_key, payload[json_key]))

    if not updates:
        return jsonify({'error': 'no editable fields provided'}), 400

    updates.append('last_edited_at = ?')
    params.append(datetime.datetime.now().isoformat(timespec='seconds'))
    updates.append('is_human_verified = 1')
    params.append(company_id)

    sql = f'UPDATE companies SET {", ".join(updates)} WHERE "Company ID" = ?'
    db.execute(sql, params)
    db.commit()

    row = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (company_id,)).fetchone()
    if row is None:
        return jsonify({'error': 'not found'}), 404
    return jsonify(row_to_company(dict(row)))


@app.route('/api/companies', methods=['POST'])
@require_edit_access
def api_company_create():
    db = get_db()
    payload = request.get_json(force=True) or {}
    name = (payload.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Company name is required'}), 400

    import random
    existing_ids = {r[0] for r in db.execute('SELECT "Company ID" FROM companies')}
    while True:
        cid = f"MANUAL-{random.randint(1000000, 9999999)}"
        if cid not in existing_ids:
            break

    cols = ['Company ID', 'Companies']
    vals = [cid, name]
    for json_key, col in COMPANY_FIELD_MAP.items():
        if json_key == 'name' or json_key not in payload:
            continue
        cols.append(col)
        vals.append(coerce_company_value(json_key, payload[json_key]))
    cols.append('Notes')
    vals.append((payload.get('notes') or 'Manually added via the platform UI.'))
    cols.append('last_edited_at')
    vals.append(datetime.datetime.now().isoformat(timespec='seconds'))
    cols.append('is_human_verified')
    vals.append(1)

    placeholders = ','.join('?' for _ in cols)
    col_sql = ','.join(f'"{c}"' for c in cols)
    db.execute(f'INSERT INTO companies ({col_sql}) VALUES ({placeholders})', vals)
    db.commit()

    row = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (cid,)).fetchone()
    return jsonify(row_to_company(dict(row)))


@app.route('/api/companies/bulk', methods=['POST'])
@require_edit_access
def api_company_bulk_create():
    db = get_db()
    payload = request.get_json(force=True) or {}
    names = payload.get('names')
    if not isinstance(names, list) or not names:
        return jsonify({'error': 'A non-empty "names" list is required'}), 400

    import random
    existing_ids = {r[0] for r in db.execute('SELECT "Company ID" FROM companies')}
    now = datetime.datetime.now().isoformat(timespec='seconds')
    created = []
    for raw_name in names:
        name = (raw_name or '').strip()
        if not name:
            continue
        while True:
            cid = f"MANUAL-{random.randint(1000000, 9999999)}"
            if cid not in existing_ids:
                existing_ids.add(cid)
                break
        db.execute(
            'INSERT INTO companies ("Company ID", "Companies", "Notes", last_edited_at, is_human_verified) VALUES (?,?,?,?,?)',
            (cid, name, 'Manually bulk-added via the platform UI.', now, 1)
        )
        created.append(cid)
    db.commit()
    return jsonify({'ok': True, 'created': len(created), 'ids': created})


@app.route('/api/companies/<company_id>', methods=['DELETE'])
@require_edit_access
def api_company_delete(company_id):
    db = get_db()
    ensure_shortlist_table(db)
    ensure_screening_scorecards_table(db)
    ensure_custom_field_tables(db)
    db.execute('DELETE FROM companies WHERE "Company ID" = ?', (company_id,))
    db.execute('DELETE FROM shortlist WHERE company_id = ?', (company_id,))
    db.execute('DELETE FROM screening_scorecards WHERE company_id = ?', (company_id,))
    db.execute('DELETE FROM custom_field_values WHERE company_id = ?', (company_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/companies/bulk-delete', methods=['POST'])
@require_edit_access
def api_company_bulk_delete():
    db = get_db()
    ensure_shortlist_table(db)
    ensure_screening_scorecards_table(db)
    ensure_custom_field_tables(db)
    payload = request.get_json(force=True) or {}
    ids = payload.get('ids')
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'A non-empty "ids" list is required'}), 400
    placeholders = ','.join('?' for _ in ids)
    db.execute(f'DELETE FROM companies WHERE "Company ID" IN ({placeholders})', ids)
    db.execute(f'DELETE FROM shortlist WHERE company_id IN ({placeholders})', ids)
    db.execute(f'DELETE FROM screening_scorecards WHERE company_id IN ({placeholders})', ids)
    db.execute(f'DELETE FROM custom_field_values WHERE company_id IN ({placeholders})', ids)
    db.commit()
    return jsonify({'ok': True, 'deleted': len(ids)})


VALID_RANGES = {'1mo', '6mo', '1y', '5y'}


@app.route('/api/prices/<ticker>')
def api_prices(ticker):
    db = get_db()
    ensure_price_cache_table(db)

    range_param = request.args.get('range', '1y')
    if range_param not in VALID_RANGES:
        return jsonify({'error': 'invalid range'}), 400

    cached = db.execute(
        'SELECT fetched_at, data FROM price_cache WHERE ticker = ? AND range_key = ?',
        (ticker, range_param)
    ).fetchone()

    if cached:
        fetched_at = datetime.datetime.fromisoformat(cached['fetched_at'])
        age_hours = (datetime.datetime.now() - fetched_at).total_seconds() / 3600
        if age_hours < CACHE_TTL_HOURS:
            return jsonify({'points': json.loads(cached['data']), 'cached': True})

    try:
        import yfinance as yf
        hist = yf.Ticker(ticker).history(period=range_param)
        if hist.empty:
            return jsonify({'error': f'No price data found for {ticker}'}), 404
        points = [
            {'date': idx.strftime('%Y-%m-%d'), 'close': round(float(row['Close']), 2)}
            for idx, row in hist.iterrows()
        ]
    except Exception as exc:
        return jsonify({'error': f'Failed to fetch {ticker}: {exc}'}), 502

    db.execute(
        'INSERT INTO price_cache (ticker, range_key, fetched_at, data) VALUES (?, ?, ?, ?) '
        'ON CONFLICT(ticker, range_key) DO UPDATE SET fetched_at = excluded.fetched_at, data = excluded.data',
        (ticker, range_param, datetime.datetime.now().isoformat(timespec='seconds'), json.dumps(points))
    )
    db.commit()

    return jsonify({'points': points, 'cached': False})


@app.route('/api/comments')
def api_comments_list():
    db = get_db()
    ensure_comments_table(db)
    rows = db.execute('SELECT * FROM comments ORDER BY id DESC').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/comments', methods=['POST'])
@require_edit_access
def api_comments_add():
    db = get_db()
    ensure_comments_table(db)
    payload = request.get_json(force=True) or {}
    text = (payload.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'Comment text is required'}), 400
    db.execute(
        'INSERT INTO comments (author, category, text, page, created_at) VALUES (?,?,?,?,?)',
        (
            (payload.get('author') or '').strip() or 'Anonymous',
            payload.get('category', 'General'),
            text,
            payload.get('page', ''),
            datetime.datetime.now().isoformat(timespec='seconds'),
        )
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/comments/<int:comment_id>', methods=['DELETE'])
@require_edit_access
def api_comments_delete(comment_id):
    db = get_db()
    ensure_comments_table(db)
    db.execute('DELETE FROM comments WHERE id = ?', (comment_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/shortlist')
def api_shortlist_list():
    db = get_db()
    ensure_shortlist_table(db)
    rows = db.execute('SELECT * FROM shortlist ORDER BY added_at DESC').fetchall()
    out = []
    for sl in rows:
        sl = dict(sl)
        crow = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (sl['company_id'],)).fetchone()
        company = row_to_company(dict(crow)) if crow else None
        out.append({
            'companyId': sl['company_id'], 'watchlist': sl['watchlist'], 'rating': sl['rating'],
            'status': sl['status'], 'rationale': sl['rationale'], 'notes': sl['notes'],
            'addedAt': sl['added_at'], 'company': company,
        })
    return jsonify(out)


@app.route('/api/shortlist/<company_id>', methods=['POST'])
@require_edit_access
def api_shortlist_upsert(company_id):
    db = get_db()
    ensure_shortlist_table(db)
    payload = request.get_json(force=True) or {}
    existing = db.execute('SELECT company_id FROM shortlist WHERE company_id = ?', (company_id,)).fetchone()
    if existing:
        db.execute(
            'UPDATE shortlist SET watchlist = ?, rating = ?, status = ?, rationale = ?, notes = ? WHERE company_id = ?',
            (payload.get('watchlist', 'Default'), payload.get('rating'), payload.get('status', 'Under Review'),
             payload.get('rationale'), payload.get('notes'), company_id)
        )
    else:
        db.execute(
            'INSERT INTO shortlist (company_id, watchlist, rating, status, rationale, notes, added_at) VALUES (?,?,?,?,?,?,?)',
            (company_id, payload.get('watchlist', 'Default'), payload.get('rating'), payload.get('status', 'Under Review'),
             payload.get('rationale'), payload.get('notes'), datetime.datetime.now().isoformat(timespec='seconds'))
        )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/shortlist/<company_id>', methods=['DELETE'])
@require_edit_access
def api_shortlist_remove(company_id):
    db = get_db()
    ensure_shortlist_table(db)
    db.execute('DELETE FROM shortlist WHERE company_id = ?', (company_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/score-templates')
def api_score_templates_list():
    db = get_db()
    ensure_score_templates_table(db)
    rows = db.execute('SELECT * FROM score_templates ORDER BY id DESC').fetchall()
    out = []
    for r in rows:
        r = dict(r)
        r['categories'] = json.loads(r.pop('categories_json'))
        out.append(r)
    return jsonify(out)


@app.route('/api/score-templates', methods=['POST'])
@require_edit_access
def api_score_templates_add():
    db = get_db()
    ensure_score_templates_table(db)
    payload = request.get_json(force=True) or {}
    name = (payload.get('name') or '').strip()
    categories = payload.get('categories')
    if not name:
        return jsonify({'error': 'Template name is required'}), 400
    if not isinstance(categories, list) or not categories:
        return jsonify({'error': 'At least one category is required'}), 400
    db.execute(
        'INSERT INTO score_templates (name, categories_json, created_at) VALUES (?,?,?)',
        (name, json.dumps(categories), datetime.datetime.now().isoformat(timespec='seconds'))
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/score-templates/<int:template_id>', methods=['DELETE'])
@require_edit_access
def api_score_templates_delete(template_id):
    db = get_db()
    ensure_score_templates_table(db)
    db.execute('DELETE FROM score_templates WHERE id = ?', (template_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/screening-scorecards')
def api_screening_scorecards_list():
    db = get_db()
    ensure_screening_scorecards_table(db)
    rows = db.execute('SELECT * FROM screening_scorecards').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/screening-scorecards/<company_id>', methods=['POST'])
@require_edit_access
def api_screening_scorecards_upsert(company_id):
    db = get_db()
    ensure_screening_scorecards_table(db)
    payload = request.get_json(force=True) or {}
    existing = db.execute('SELECT company_id FROM screening_scorecards WHERE company_id = ?', (company_id,)).fetchone()
    fields = ('management_quality', 'integration_complexity', 'market_timing', 'deal_feasibility', 'manually_added')
    values = [payload.get(f) for f in fields]
    now = datetime.datetime.now().isoformat(timespec='seconds')
    if existing:
        db.execute(
            'UPDATE screening_scorecards SET management_quality=?, integration_complexity=?, market_timing=?, '
            'deal_feasibility=?, manually_added=?, updated_at=? WHERE company_id=?',
            (*values, now, company_id)
        )
    else:
        db.execute(
            'INSERT INTO screening_scorecards (company_id, management_quality, integration_complexity, '
            'market_timing, deal_feasibility, manually_added, updated_at) VALUES (?,?,?,?,?,?,?)',
            (company_id, *values, now)
        )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/screening-scorecards/<company_id>', methods=['DELETE'])
@require_edit_access
def api_screening_scorecards_remove(company_id):
    db = get_db()
    ensure_screening_scorecards_table(db)
    db.execute('DELETE FROM screening_scorecards WHERE company_id = ?', (company_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/custom-fields')
def api_custom_fields_list():
    db = get_db()
    ensure_custom_field_tables(db)
    rows = db.execute('SELECT * FROM custom_field_defs ORDER BY id').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/custom-fields', methods=['POST'])
@require_edit_access
def api_custom_fields_create():
    db = get_db()
    ensure_custom_field_tables(db)
    payload = request.get_json(force=True) or {}
    name = (payload.get('name') or '').strip()
    field_type = (payload.get('field_type') or 'text').strip()
    if not name:
        return jsonify({'error': 'Field name is required'}), 400
    if field_type not in ('text', 'number', 'date', 'tag'):
        field_type = 'text'
    existing = db.execute('SELECT id FROM custom_field_defs WHERE name = ?', (name,)).fetchone()
    if existing:
        return jsonify({'error': 'A custom field with that name already exists'}), 400
    db.execute(
        'INSERT INTO custom_field_defs (name, field_type, created_at) VALUES (?,?,?)',
        (name, field_type, datetime.datetime.now().isoformat(timespec='seconds'))
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/custom-fields/<int:field_id>', methods=['DELETE'])
@require_edit_access
def api_custom_fields_delete(field_id):
    db = get_db()
    ensure_custom_field_tables(db)
    db.execute('DELETE FROM custom_field_defs WHERE id = ?', (field_id,))
    db.execute('DELETE FROM custom_field_values WHERE field_id = ?', (field_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/custom-fields/values/<company_id>', methods=['POST'])
@require_edit_access
def api_custom_field_values_upsert(company_id):
    db = get_db()
    ensure_custom_field_tables(db)
    payload = request.get_json(force=True) or {}
    field_id = payload.get('field_id')
    value = payload.get('value')
    if field_id is None:
        return jsonify({'error': 'field_id is required'}), 400
    existing = db.execute(
        'SELECT 1 FROM custom_field_values WHERE company_id = ? AND field_id = ?', (company_id, field_id)
    ).fetchone()
    if existing:
        db.execute(
            'UPDATE custom_field_values SET value = ? WHERE company_id = ? AND field_id = ?',
            (value, company_id, field_id)
        )
    else:
        db.execute(
            'INSERT INTO custom_field_values (company_id, field_id, value) VALUES (?,?,?)',
            (company_id, field_id, value)
        )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/dashboard-widgets')
def api_dashboard_widgets_list():
    db = get_db()
    ensure_dashboard_widgets_table(db)
    rows = db.execute('SELECT * FROM dashboard_widgets ORDER BY position, id').fetchall()
    out = []
    for r in rows:
        r = dict(r)
        r['config'] = json.loads(r.pop('config_json'))
        out.append(r)
    return jsonify(out)


@app.route('/api/dashboard-widgets', methods=['POST'])
@require_edit_access
def api_dashboard_widgets_create():
    db = get_db()
    ensure_dashboard_widgets_table(db)
    payload = request.get_json(force=True) or {}
    wtype = (payload.get('type') or '').strip()
    title = (payload.get('title') or '').strip()
    config = payload.get('config') or {}
    if wtype not in ('kpi', 'bargroup', 'topn'):
        return jsonify({'error': 'Invalid widget type'}), 400
    if not title:
        return jsonify({'error': 'Widget title is required'}), 400
    max_pos = db.execute('SELECT COALESCE(MAX(position), -1) FROM dashboard_widgets').fetchone()[0]
    db.execute(
        'INSERT INTO dashboard_widgets (type, title, config_json, position, created_at) VALUES (?,?,?,?,?)',
        (wtype, title, json.dumps(config), max_pos + 1, datetime.datetime.now().isoformat(timespec='seconds'))
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/dashboard-widgets/<int:widget_id>', methods=['DELETE'])
@require_edit_access
def api_dashboard_widgets_delete(widget_id):
    db = get_db()
    ensure_dashboard_widgets_table(db)
    db.execute('DELETE FROM dashboard_widgets WHERE id = ?', (widget_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/dashboard-widgets/<int:widget_id>/position', methods=['POST'])
@require_edit_access
def api_dashboard_widgets_reposition(widget_id):
    db = get_db()
    ensure_dashboard_widgets_table(db)
    payload = request.get_json(force=True) or {}
    position = payload.get('position', 0)
    db.execute('UPDATE dashboard_widgets SET position = ? WHERE id = ?', (position, widget_id))
    db.commit()
    return jsonify({'ok': True})


FRIENDLY_COL_NAMES = {
    'id': 'Company ID', 'name': 'Company', 'website': 'Website', 'description': 'Description',
    'segment': 'Primary Segment', 'segment2': 'Secondary Segment', 'businessModel': 'Business Model',
    'geo': 'Geography Region', 'hq': 'HQ Location', 'yearFounded': 'Year Founded',
    'capitalIntensity': 'Capital Intensity', 'notes': 'Notes', 'revenueModel': 'Revenue Model',
    'targetCustomer': 'Target Customer', 'gtm': 'Go-to-Market Motion', 'moat': 'Moat',
    'strategyNotes': 'Strategy Notes', 'ticker': 'Ticker', 'tickerSource': 'Ticker Source',
    'tickerNote': 'Ticker Note', 'totalRaised': 'Total Raised ($M)', 'ownershipStatus': 'Ownership Status',
    'businessStatus': 'Business Status', 'activeInvestors': 'Active Investors', 'competitors': 'Competitors',
    'lastEditedAt': 'Last Edited At', 'isHumanVerified': 'Human Verified',
    'lastFinancingDate': 'Last Financing Date', 'lastFinancingSize': 'Last Financing Size ($M)',
    'lastFinancingValuation': 'Last Financing Valuation ($M)', 'valuationEstimate': 'Valuation Estimate ($M)',
    'lastFinancingDealType': 'Last Financing Deal Type', 'companyFinancingStatus': 'Financing Status',
    'country': 'Country', 'parentCompany': 'Parent Company', 'primaryContact': 'Primary Contact',
    'successProbability': 'Success Probability', 'primaryIndustryCode': 'PitchBook Industry Code',
    'emergingSpaces': 'Emerging Spaces',
    'verticals': 'Verticals',
}


@app.route('/api/export/companies.xlsx')
def export_companies_xlsx():
    import pandas as pd
    from flask import send_file
    from openpyxl import Workbook
    from excel_format import style_workbook
    import io

    ids_param = request.args.get('ids', '')
    db = get_db()
    if ids_param:
        ids = [i for i in ids_param.split(',') if i]
        rows = []
        for cid in ids:
            r = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (cid,)).fetchone()
            if r:
                rows.append(row_to_company(dict(r)))
    else:
        rows = [row_to_company(dict(r)) for r in db.execute('SELECT * FROM companies').fetchall()]

    df = pd.DataFrame(rows)
    df = df.rename(columns=FRIENDLY_COL_NAMES)

    wb = Workbook()
    ws = wb.active
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))
    style_workbook(ws, df, 'Companies')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='companies_export.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/shortlist.xlsx')
def export_shortlist_xlsx():
    import pandas as pd
    from flask import send_file
    from openpyxl import Workbook
    from excel_format import style_workbook
    import io

    db = get_db()
    ensure_shortlist_table(db)
    sl_rows = db.execute('SELECT * FROM shortlist ORDER BY added_at DESC').fetchall()
    records = []
    for sl in sl_rows:
        sl = dict(sl)
        crow = db.execute('SELECT * FROM companies WHERE "Company ID" = ?', (sl['company_id'],)).fetchone()
        c = row_to_company(dict(crow)) if crow else {}
        records.append({
            'Company': c.get('name'), 'Watchlist': sl['watchlist'], 'Rating': sl['rating'], 'Status': sl['status'],
            'Rationale': sl['rationale'], 'Notes': sl['notes'], 'Added At': sl['added_at'],
            'Segment': c.get('segment'), 'Business Model': c.get('businessModel'), 'Geography': c.get('geo'),
            'Country': c.get('country'), 'Total Raised ($M)': c.get('totalRaised'),
            'Valuation ($M)': c.get('lastFinancingValuation') or c.get('valuationEstimate'),
            'Moat': c.get('moat'), 'Website': c.get('website'), 'Description': c.get('description'),
        })

    df = pd.DataFrame(records)
    wb = Workbook()
    ws = wb.active
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=(None if pd.isna(val) else val))
    style_workbook(ws, df, 'Shortlist')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='investment_shortlist.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/prypco')
def api_prypco():
    global _PRYPCO_CACHE
    if _PRYPCO_CACHE is None:
        _PRYPCO_CACHE = load_prypco_data()
    return jsonify(_PRYPCO_CACHE)


@app.route('/api/meta')
def api_meta():
    return jsonify({
        'segments': [
            'Distribution & Sales', 'Underwriting & Risk Assessment', 'Policy Administration & Core Systems',
            'Claims Management', 'Customer & Member Engagement', 'Full-Stack Carriers / MGAs', 'Reinsurance Tech',
            'Compliance / RegTech', 'Benefits Administration', 'Infrastructure / Enabling Tech', 'Adjacent / Non-Core'
        ],
        'models': [
            'Software / SaaS', 'Full-Stack Carrier', 'MGA / Delegated Underwriting Authority', 'Broker / Agency',
            'Marketplace / Comparison Platform', 'Data, Analytics & API Provider',
            'Embedded / Affinity Insurance Enabler', 'Services / BPO'
        ],
        'geos': ['North America', 'Europe', 'Asia-Pacific', 'Middle East & Africa', 'Latin America', 'Unknown'],
        'capitalOptions': ['Capital-heavy (balance-sheet risk)', 'Capital-light (software/services)'],
        'revenueModels': [
            'Subscription / License Fees', 'Commission / Take Rate', 'Transaction / Usage Fees',
            'Premium Underwriting Spread', 'Services / Project Fees', 'Data Licensing',
            'Freemium / Ad-Supported', 'Mixed / Unclear'
        ],
        'targetCustomers': [
            'B2B – Insurers & Reinsurers', 'B2B – Brokers & Agents', 'B2B – Enterprise / Employer',
            'B2B2C – Embedded via Partner', 'B2C – Direct to Consumer'
        ],
        'gtmMotions': [
            'Direct Sales', 'Channel / Broker Partnerships', 'Embedded / API Integration',
            'Comparison / Marketplace Acquisition', 'Affinity / Co-Brand Partnerships',
            'Performance Marketing / Direct-to-Consumer'
        ],
        'moats': [
            'Proprietary Data & Analytics', 'Regulatory License (MGA/Carrier Authority)',
            'Network Effects (Marketplace Liquidity)', 'Switching Costs / Deep System Integration',
            'Brand & Distribution Scale', 'Low Differentiation (Commodity Software)'
        ],
    })


if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG', '1') == '1', port=int(os.environ.get('PORT', 8080)), host='0.0.0.0')
